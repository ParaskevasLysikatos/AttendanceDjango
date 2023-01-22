# the seeders take the data from excel files and fill the tables with data,attention the excel should be properly formed with data in corresponding row-cell,and declared in exact file path in media folder,enable them in urls.py
from django.shortcuts import render,redirect
from django.http import HttpResponseRedirect,HttpResponse,JsonResponse
from django.template import loader
import sys,os,logging,datetime,json
sys.path.insert(1,os.path.abspath("./pyzk"))
from zk import ZK, const
from openpyxl import load_workbook

excelSeed = "ipearanceFront/media/usersSeedV3.xlsx"
excelNameDevices = "ipearanceFront/media/devices.xlsx"
excelDep="ipearanceFront/media/department.xlsx"
excelStaffDep="ipearanceFront/media/staff_dep_info.xlsx"
roles="ipearanceFront/media/roles.xlsx"
industry="ipearanceFront/media/industry.xlsx"
members="ipearanceFront/media/members.xlsx"
relationT="ipearanceFront/media/relationTypes.xlsx"

from django.utils.dateparse import parse_datetime,parse_date,parse_time
from ipearanceBackend.models import Users,Devices,Roles,Departments,Industry,Staff_department_info,Signatures,Members_team,Relation_types 
                
from ipearanceBackend.serializers import UsersSerializer,DevicesSerializer,DepartmentsSerializer,RolesSerializer,IndustrySerializer,Staff_department_infoSerializer,SignaturesSerializer,MembersTeamSerializer,Relation_typesSerializer
logger = logging.getLogger(__name__)
excel = list()

def seedUsers(request):
    wb = load_workbook(excelSeed)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        logger.error(row_data)
        logger.error(row_data[0])
        u=Users(
            id=int(row_data[0]),
            device_card_id=int(row_data[8]),
            device_card_number=int(row_data[10]),
            username=(row_data[9]),
            first_name=(row_data[1].replace("'","")),
            last_name=(row_data[2].replace("'","")),
			fathers_name="--",#needed for front searches
            email=(row_data[6].replace("'","")),
            hrms_id=int(row_data[11]),
            date_joined=datetime.datetime.now(),
            is_active=True,
            is_staff=True,
            member_team_name=int(row_data[12])
             )
        u.set_password(row_data[7])
        u.save()
        logger.error(row_data[0])
    allUsers=list(Users.objects.all())
    data=UsersSerializer(allUsers,many=True)
    return HttpResponse(data,content_type="application/json")


    
def seedDevices(request):
    wb = load_workbook(excelNameDevices)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        logger.error(row_data)
        logger.error(row_data[0])
        d=Devices(
            device_location=row_data[0],
            device_ip=row_data[1],
            device_port=4370
             )
        d.save()
    allDevices=list(Devices.objects.all())
    data=DevicesSerializer(allDevices,many=True)
    return HttpResponse(data,content_type="application/json") 
    
def seedDepartments(request):
    wb = load_workbook(excelDep)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        logger.error(row_data)
        logger.error(row_data[0])
        d=Departments(
            id=int(row_data[0]),
            industry_id = row_data[1] ,
            parent_id = row_data[2],
            department_name = row_data[3],
             )
        d.save()
    allDepartments=list(Departments.objects.all())
    data=DepartmentsSerializer(allDepartments,many=True)
    return HttpResponse(data,content_type="application/json") 
    
def seedRoles(request):
    wb = load_workbook(roles)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        logger.error(row_data)
        logger.error(row_data[0])
        sr=Roles(
            role_name = row_data[0]
             )
        sr.save()
    allRoles=list(Roles.objects.all())
    data=RolesSerializer(allRoles,many=True)
    return HttpResponse(data,content_type="application/json")
    
def seedRelationTypes(request):
    wb = load_workbook(relationT)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        logger.error(row_data)
        logger.error(row_data[0])
        rt=Relation_types(
            relation_type_name = row_data[0]
             )
        rt.save()
    allRelations=list(Relation_types.objects.all())
    data=Relation_typesSerializer(allRelations,many=True)
    return HttpResponse(data,content_type="application/json")
    
def seedMembers(request):
    wb = load_workbook(members)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        logger.error(row_data)
        logger.error(row_data[0])
        m=Members_team(
            member_team_name = row_data[0].strip(),
            member_team_weight = row_data[1]
             )
        m.save()
    allMembers=list(Members_team.objects.all())
    data=MembersTeamSerializer(allMembers,many=True)
    return HttpResponse(data,content_type="application/json")
    
def seedSignatures(request):
    wb = load_workbook(roles)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        logger.error(row_data)
        logger.error(row_data[0])
        s=Signatures(
            signature_name = row_data[0]
             )
        s.save()
    allSignatures=list(Signatures.objects.all())
    data=SignaturesSerializer(allSignatures,many=True)
    return HttpResponse(data,content_type="application/json")
       
def seedIndustry(request):
    wb = load_workbook(industry)
    worksheet = wb["Sheet1"]
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        logger.error(row_data)
        logger.error(row_data[0])
        i=Industry(
            name = row_data[0],
            address=row_data[1],
            afm=int(row_data[2]),
            hrms_connect=row_data[3],
            hrms_ip =row_data[4],
            hrms_port =int( row_data[5]),
            hrms_service_name = row_data[6],
            hrms_username = row_data[7],
            hrms_password = row_data[8],
            hrms_schema='',
            hrms_com_id= int(row_data[9])
             )
        i.save()
    allIndustry=list(Industry.objects.all())
    data=IndustrySerializer(allIndustry,many=True)
    return HttpResponse(data,content_type="application/json") 
    
def seedStaffDepartments(request):
    wb = load_workbook(excelStaffDep)
    worksheet = wb["Sheet1"]
    roleDict = {
            3:"ΥΠΑΛΛΗΛΟΣ",
            2:"ΠΡΟΪΣΤΑΜΕΝΟΣ",
            1:"ΔΙΕΥΘΥΝΤΗΣ"
            }
    dict = {
            0:"False",
            1:"True"
            }
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        d=Staff_department_info(
            department_id = row_data[1] ,
            #staff_id = row_data[2],
            staff_card=Users.objects.get(device_card_id=int(row_data[2])).device_card_number,
            role = roleDict.get(int(row_data[3])),
            
            apply_from = parse_datetime(row_data[4]) ,
            apply_to = parse_datetime(row_data[5]),
            
            start_of_work = parse_time(row_data[6]),
            end_of_work = parse_time(row_data[7]),
            Monday=dict.get(int(row_data[8])),
            Tuesday=dict.get(int(row_data[9])),
            Wednesday=dict.get(int(row_data[10])),
            Thursday=dict.get(int(row_data[11])),
            Friday=dict.get(int(row_data[12])),
            Saturday=dict.get(int(row_data[13])),
            Sunday=dict.get(int(row_data[14]))
             )
        d.save()
    allStaffDepartmentInfo=list(Staff_department_info.objects.all())
    data=Staff_department_infoSerializer(allStaffDepartmentInfo,many=True)
    return HttpResponse(data,content_type="application/json") 
    





        


